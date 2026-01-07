import openpyxl
import os
import sys

# Configuration
input_file = r'c:\Users\001\Desktop\14-Relation\03-Handmade\01-Handmade.xlsx'
output_file = r'c:\Users\001\Desktop\14-Relation\03-Handmade\01-Handmade_filtered.xlsx'

def is_valid_qid(value):
    if value is None:
        return False
    s = str(value).strip()
    return len(s) > 0 and s.lower() != 'none' and s.lower() != 'nan'

def process():
    print("Starting process...", flush=True)
    if not os.path.exists(input_file):
        print(f"Error: Input file not found: {input_file}", flush=True)
        return

    print(f"Reading from: {input_file}", flush=True)
    wb = openpyxl.load_workbook(input_file, read_only=True, data_only=True)
    ws = wb.active

    # Create output workbook
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Filtered Data"

    header = []
    qid_indices = []
    
    total_rows = 0
    kept_rows = 0
    
    # Statistics
    stats = {
        'both_qids': 0,
        'only_qid1': 0,
        'only_qid2': 0,
        'no_qids': 0
    }

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            header = row
            out_ws.append(header)
            # Identify QID columns
            for idx, col_name in enumerate(row):
                if str(col_name).strip().upper() == 'QID':
                    qid_indices.append(idx)
            print(f"Found QID columns at indices: {qid_indices}", flush=True)
            if not qid_indices:
                print("Warning: No columns named 'QID' found. Exiting.", flush=True)
                return
            continue

        total_rows += 1
        
        has_qid1 = False
        has_qid2 = False
        
        # Check first QID column (Entity 1)
        if len(qid_indices) > 0:
            if is_valid_qid(row[qid_indices[0]]):
                has_qid1 = True
                
        # Check second QID column (Entity 2)
        if len(qid_indices) > 1:
            if is_valid_qid(row[qid_indices[1]]):
                has_qid2 = True
        
        # Update stats
        if has_qid1 and has_qid2:
            stats['both_qids'] += 1
        elif has_qid1:
            stats['only_qid1'] += 1
        elif has_qid2:
            stats['only_qid2'] += 1
        else:
            stats['no_qids'] += 1

        # Filter Logic: Keep if ANY QID is present
        if has_qid1 or has_qid2:
            out_ws.append(row)
            kept_rows += 1

    print("-" * 30, flush=True)
    print(f"Processing Complete.", flush=True)
    print(f"Total data rows processed: {total_rows}", flush=True)
    print(f"Rows with BOTH QIDs: {stats['both_qids']}", flush=True)
    print(f"Rows with ONLY Entity 1 QID: {stats['only_qid1']}", flush=True)
    print(f"Rows with ONLY Entity 2 QID: {stats['only_qid2']}", flush=True)
    print(f"Rows with NO QIDs: {stats['no_qids']}", flush=True)
    print("-" * 30, flush=True)
    print(f"Total rows kept (Any QID): {kept_rows}", flush=True)
    print(f"Saving to: {output_file}", flush=True)
    
    out_wb.save(output_file)
    print("Done.", flush=True)

if __name__ == "__main__":
    process()
