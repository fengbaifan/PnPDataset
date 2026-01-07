import openpyxl
import os
import sys

file_path = r'c:\Users\001\Desktop\14-Relation\03-Handmade\01-Handmade.xlsx'

print(f"Checking file: {file_path}")
sys.stdout.flush()

if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    exit(1)

try:
    print("Loading workbook...")
    sys.stdout.flush()
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    print(f"Workbook loaded. Sheet names: {wb.sheetnames}")
    ws = wb.active
    print(f"Active sheet: {ws.title}")
    sys.stdout.flush()
    
    # Print the first few rows to understand structure
    count = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        print(f"Row {i+1}: {row}")
        sys.stdout.flush()
        count += 1
        if count >= 5: # Print first 5 rows
            break
    
    if count == 0:
        print("Sheet is empty.")
            
except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
